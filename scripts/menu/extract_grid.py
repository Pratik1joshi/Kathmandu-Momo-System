#!/usr/bin/env python3
"""
Faithful xlsx -> JSON grid extractor for the Dim Sum Puri menu workbook.

This does NO interpretation: it dumps every cell of the source sheet as a 2D
grid (with 1-based row/column coordinates preserved) so the Node importer can
apply category/variant logic while every price remains traceable to a cell.

Usage:
  python scripts/menu/extract_grid.py                # default paths
  python scripts/menu/extract_grid.py in.xlsx out.json

Output: data/menu/menu-grid.json
"""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl is required: pip install openpyxl\n")
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[2]
SRC = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "data" / "menu" / "Menu.xlsx"
OUT = Path(sys.argv[2]) if len(sys.argv) > 2 else ROOT / "data" / "menu" / "menu-grid.json"


def col_letter(idx0):
    """0-based column index -> spreadsheet letter (A, B, ...)."""
    idx = idx0 + 1
    s = ""
    while idx:
        idx, rem = divmod(idx - 1, 26)
        s = chr(65 + rem) + s
    return s


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.worksheets[0]
    grid = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = []
        for c_idx, val in enumerate(row):
            if isinstance(val, str):
                val = val.strip()
                if val == "":
                    val = None
            cells.append(val)
        grid.append(cells)

    payload = {
        "source_file": str(SRC.relative_to(ROOT)).replace("\\", "/"),
        "sheet": ws.title,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "col_letters": [col_letter(i) for i in range(ws.max_column)],
        "grid": grid,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    non_empty = sum(1 for row in grid for c in row if c is not None)
    print(f"Wrote {OUT.relative_to(ROOT)} ({len(grid)} rows, {non_empty} non-empty cells)")


if __name__ == "__main__":
    main()
